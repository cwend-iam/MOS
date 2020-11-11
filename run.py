# Libraries
import datetime
import os
import re
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Global variabelen
input_folder = "Siemens planning"

output_folder = "MOS planning"

export_df_format = pd.DataFrame({'Job nummer': [],
                                 'Job omschrijving': [],
                                 'Uitvoerende': [],
                                 'PO': [],
                                 'Frequentie': [],
                                 'Begin datum': [],
                                 'Eind datum': [],
                                 'SB': [],
                                 'HP': [],
                                 'TRA': [],
                                 'SPP': [],
                                 'Object': [],
                                 'Aannemer': [],
                                 'Wie': [],
                                 'Telefoonnummer': []})

max_week_num = 53

# Kleuren
"""
De pd.Series kleuren_series is een serie waarin per regel van export_df wordt bijgehouden welke kleur 
de kolom Job nummer heeft. bij het schrijven van de kleuren in het export document worden beide parallel 
aan elkaar doorlopen.
De pd.Series kleuren_series wordt opgevuld met de volgende waarden:
    -   0 - geen kleur
    -   1 - geel
    -   2 - groen
De kleurencodes zijn hieronder bij de andere kleuren gedefinieerd.
"""
empty_tab_fill = 'FFFF0800'
white = 'FFFFFFFF'
black = '00000000'
orange = 'FFFF6600'

COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',  # 5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
    '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
    '00993300', '00993366', '00333399', '00333333',  # 60-63
)

no_fill = PatternFill(fill_type=None)
fill = PatternFill(fill_type='solid', start_color=orange, end_color=orange)


# Definiëren van standaard functies
def import_workbook(file):
    """
    Het importeren van de workbook aan de hand van een window waarin de gebruiker zelf kan selecteren welk bestand
    zij geïmporteerd willen hebben.
    :param file: Het bestand dat men wilt importeren.
    :return: Een workbook of een foutmelding
    """
    # todo: print statements aanpassen zodat deze correct werkt met de nieuwe vorm van toepassen van in- en outputs.
    if file.lower().endswith('.xlsm') or file.lower().endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(filename=file, read_only=False)
            return workbook
        except:
            print('Er fout opgetreden tijdens het lezen van het door u geselecteerde bestand.')
            return sys.exit()
    else:
        print('Het door u geselecteerde bestand kan niet gelezen worden.')
        return sys.exit()


def week_rijen(tabblad, week_nummer):
    """
    Een fucntie specifiek gemaakt voor de Siemens planning. De Functie bepaalt hoeveel regels een bepaalde
    week in neemt.
    In de terugkoppeling is het rijnummer het nummer van de eerste rij data van de gevraagde week.
    :param tabblad: Sheet van de workbook (siemens planning)
    :param week_nummer: De week waarvan men het aantal regels wilt achterhalen.
    :return: Een tuple(startrij, eindrij)
    """

    global max_week_num

    # Controleren of het de laatste week van het jaar is
    if week_nummer == max_week_num:
        # zoeken naar start en eind rij voor laatste week in document
        for row in range(1, tabblad.max_row + 1):

            if tabblad[f'B{row}'].value == 'Week':

                if tabblad[f'C{row}'].value == week_nummer:
                    start = row + 1
                else:
                    pass

            if tabblad[f'B{row}'].value is None and tabblad[f'C{row}'].value is None:
                eind = row

        return start, eind

    # Else-statement voor wanneer het niet de laatste week van het jaar is
    else:
        for row in range(1, tabblad.max_row + 1):
            if tabblad[f'B{row}'].value == 'Week':

                if tabblad[f'C{row}'].value == week_nummer:
                    start = row + 1
                else:
                    pass

                if tabblad[f'C{row}'].value == (week_nummer + 1):
                    eind = row
                    break
                else:
                    pass

        return start, eind


def set_tab_color(workbook):
    global black, white, empty_tab_fill

    for sheet in workbook.worksheets:
        if sheet.max_row > 1:
            sheet.sheet_properties.tabColor = black
        else:
            sheet.sheet_properties.tabColor = empty_tab_fill


if __name__ == '__main__':
    # todo: De code aanpassen zodat de werkmap intuïtief werkt
    # Bericht voor de gebruiker
    print(f'Deze automatisering is gemaakt voor het overschrijven van gegevens uit de Siemens planning naar '
          f'de SHEAPA planning. ')

    # Genereren van een lijst met de input bestanden
    _list_input_files = os.listdir(input_folder)

    list_input_files = list()

    for f in _list_input_files:
        if f == '.gitignore':
            pass
        else:
            list_input_files.append(f)

    count = int()
    total = len(list_input_files)

    if total == 0:
        print("MELDIING: De map \'Siemens planning\' is leeg. Upload een Siemens planning naar deze map en start het "
              "programma opnieuw.")
        sys.exit()
    else:
        pass

    for input_file in list_input_files:
        count += 1

        # Bericht voor de gebruiker
        print(f'Het bestand \'{input_file}\' wordt geïmporteerd. \n'
              f'Bestand {count} van de {total}. Even geduld AUB.')

        input_file_path = os.path.join(input_folder, input_file)

        # bestand_naam = '.\\res\\dev test obj week 38.xlsx'
        siemens_planning = import_workbook(file=input_file_path)

        # Bericht voor de gebruiker
        print(f'De gegevens worden opgehaald en verwerkt. ')

        mkn_sheet = siemens_planning['MKN']
        mkz_sheet = siemens_planning['MKZ']
        hk_sheet = siemens_planning['HK']

        sheets = [mkn_sheet, mkz_sheet, hk_sheet]

        # Ophalen van bepalende variabelen
        # Weeknummer
        if re.search(r'\bweek\b', input_file):
            start_index = re.search(r'\bweek\b', input_file).start()
            if re.search(r'(?<=\bweek\b.)\d\d', input_file):
                week_num = int(input_file[start_index + 5: start_index + 7].strip(r'.xlsx'))
            else:
                week_num = int(input_file[start_index + 5: start_index + 6])

        # Lijst met de weeknummers van de drie weken die gelezen moeten worden (ze plannen drie weken vooruit)
        te_lezen_weken = [week_num, week_num + 1, week_num + 2]

        # Verwijderen van onmogelijke weeknummers (>53)
        for i in reversed(range(len(te_lezen_weken))):
            week = te_lezen_weken[i]
            if week > max_week_num:
                te_lezen_weken.pop(i)
            else:
                pass

        # Itereren over de drie wekend die gelezen moeten worden
        for i in range(len(te_lezen_weken)):
            # De nieuwe iteratie beginnen met een leeg export dataframe
            export_df = export_df_format

            # De nieuwe iteratie beginnen met een lege pd.Series voor het vastleggen van de kleuren
            kleuren_series = pd.Series(data=[], name='kleur')

            # Ophalen van de gegevens
            for sheet in sheets:
                # Aantal rijen van een week in de siemens planning
                rijen_week = week_rijen(sheet, week_nummer=te_lezen_weken[i])

                for r in range(rijen_week[0], rijen_week[1]):
                    uitvoerende = sheet[f'H{r}'].value

                    if uitvoerende is not None and uitvoerende != ' ':
                        job_nummer = str(sheet[f'A{r}'].value)
                        job_omschrijving = sheet[f'B{r}'].value
                        po = sheet[f'I{r}'].value
                        frequentie = sheet[f'J{r}'].value
                        begin_datum = sheet[f'L{r}'].value
                        eind_datum = sheet[f'M{r}'].value
                        sb = sheet[f'N{r}'].value
                        hd = sheet[f'O{r}'].value
                        tra = sheet[f'P{r}'].value
                        ssp = sheet[f'Q{r}'].value
                        obj = sheet.title
                        aannemer = str()
                        wie = str()
                        telefoonnummer = str()

                        jobnummer_kleur = sheet[f'A{r}'].fill.fgColor

                        # Kleur van cel uitlezen en vastleggen in aparte pd.Series
                        if type(jobnummer_kleur.indexed) is int:
                            add = pd.Series(COLOR_INDEX[jobnummer_kleur.indexed])
                            kleuren_series = kleuren_series.append(add, ignore_index=True)
                        else:
                            add = pd.Series([black])
                            kleuren_series = kleuren_series.append(add, ignore_index=True)

                        # Aanpassen van het format van de datum
                        if isinstance(begin_datum, datetime.datetime):
                            begin_datum = begin_datum.strftime('%a %d-%m-%Y')
                        else:
                            pass

                        if isinstance(eind_datum, datetime.datetime):
                            eind_datum = eind_datum.strftime('%a %d-%m-%Y')
                        else:
                            pass

                        rij = pd.Series([str(job_nummer), str(job_omschrijving), str(uitvoerende), str(po),
                                         str(frequentie), str(begin_datum), str(eind_datum), str(sb), str(hd),
                                         str(tra), str(ssp), str(obj), aannemer, wie, telefoonnummer],
                                        index=export_df.columns)

                        export_df = export_df.append(rij, ignore_index=True)

                    # Als er geen uitvoeder bij staat, is de regel niet belangrijk (aka overslaan)
                    else:
                        pass

            # Verwijderen van de 'None' waarden uit de dataframe
            export_df = export_df.replace(to_replace='None', value='')

            # Genereren van een dataframe per week (elke week is een nieuw tabblad)
            if i == 0:
                df_week_1 = export_df
                kleuren_week_1 = kleuren_series
            elif i == 1:
                df_week_2 = export_df
                kleuren_week_2 = kleuren_series
            else:
                df_week_3 = export_df
                kleuren_week_3 = kleuren_series

        # Controle slag (??)
        """
        Controleren of het bestand aanwezig is. zo niet, dan moet het gehele bestand (met alle weken)
        gegenereerd worden.
        Controleren of het tabblad waar een df naar geschreven moet worden leeg is of niet, wanneer
        deze niet leeg is, moet er iets gebeuren met de regels die al in het document staan (kleur veranderen of zo)
        Of er eerst iets moet gebeuren met de oude of de nieuwe regels, kan bepaald worden wanneer het concreet
        gebouwd gaat worden. Maar de nieuwe regels oeten worden toegevoegd aan het tabblad met de correcte week.
        Eindstand opslaag en afsluiten.
        """
        # Opstellen van een lijst bestanden aanwezig in de root directory
        output_dir_files = os.listdir(output_folder)

        # Huidige jaar ophalen
        jaar = datetime.date.today().year

        # Naam van het export bestand definiëren
        export_file_name = f'Planning Jobs + Capaciteit {jaar}.xlsx'

        # Pad naar de exportlocatie definiëren
        export_file_path = os.path.join(output_folder, export_file_name)

        # Export locatie checken op de naam van het bestand
        # Elk jaar moet er een nieuw document autimatisch gegenereerd worden
        if export_file_name not in output_dir_files:
            export_wb = openpyxl.Workbook()

            # Genereren van de tabbladen
            for i in range(max_week_num + 1):
                if i == 0:
                    export_wb.create_sheet("Legenda", 0)
                    export_wb.remove(export_wb['Sheet'])
                else:
                    export_wb.create_sheet(f'Week {i}')

            for sheet in export_wb.sheetnames:
                ws = export_wb[sheet]
                if sheet == 'Legenda':
                    # Toelichtingen voor in de legenda
                    toelichting_rood = 'Een tabblad met deze kleur duidt op een leeg tabblad'
                    toelichting_zwart = 'Een tabblad met deze kleur duidt op een gevuld tabblad'
                    toelichting_oranje = 'Een regel met een job nummer in deze kleur duidt op een unieke regel die ' \
                                         'niet afkomstig is uit de eigen weekplanning van die week.'
                    toelichting_groen = 'Een regel met een job nummer in deze kleur duidt op een regel die nieuw is in' \
                                        ' de planning.'
                    toelichting_geel = 'Een regel met een job nummer in deze kleur duidt op een regel die is herpland.'
                    toelichting_012 = 'De kolommen \'SB\', \'HP\', \'TRA\', en \'SPP\' bevatten in het oorspronkelijke' \
                                      ' Siemens document pictogrammen. Deze pictogrammen zijn afhankelijk van de ' \
                                      'getallen 0, 1, en 2. In dit document wordt de cijfermatige verwijzing toegepast.'
                    toelichting_0 = 'Noodzakelijke actie, nog niet ingediend'
                    toelichting_1 = 'Ingeleverd, nog niet akkoord'
                    toelichting_2 = 'Ingeleverd en goedgekeurd door RWS'

                    # Mergen van de cellen (layout)
                    ws.merge_cells('B2:D4')  # voor titel
                    ws.merge_cells('D6:M6')  # voor toelichting kleur rood
                    ws.merge_cells('D7:M7')  # voor toelichting kleur zwart
                    ws.merge_cells('D9:M10')  # voor toelichting kleur oranje
                    ws.merge_cells('D11:M12')  # voor toelichting kleur geel
                    ws.merge_cells('D13:M14')  # voor toelichting kleur groen
                    ws.merge_cells('D16:M18')  # voor toelichting cijferschaal
                    ws.merge_cells('D20:M20')  # voor toelichting cijfer 0
                    ws.merge_cells('D21:M21')  # voor toelichting cijfer 1
                    ws.merge_cells('D22:M22')  # voor toelichting cijfer 2

                    # Inhoud titel cell (B2)
                    ws['B2'].value = 'Legenda'
                    ws['B2'].font = Font(size=26)
                    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud B6
                    ws['B6'].fill = PatternFill(fill_type='solid', start_color=empty_tab_fill, end_color=empty_tab_fill)

                    # Inhoud B7
                    ws['B7'].fill = PatternFill(fill_type='solid', start_color=black, end_color=black)

                    # Inhoud B9
                    ws['B9'].fill = PatternFill(fill_type='solid', start_color=orange, end_color=orange)

                    # Inhoud B11
                    ws['B11'].fill = PatternFill(fill_type='solid', start_color=COLOR_INDEX[34],
                                                 end_color=COLOR_INDEX[34])

                    # Inhoud B13
                    ws['B13'].fill = PatternFill(fill_type='solid', start_color=COLOR_INDEX[11],
                                                 end_color=COLOR_INDEX[11])

                    # Inhoud B16
                    c = 'B16'
                    ws[c].value = '0, 1, 2'
                    ws[c].alignment = Alignment(horizontal='left', vertical='top')

                    # Inhoud B19
                    c = 'B19'
                    ws[c].value = 'Schaalverdeling:'
                    ws[c].alignment = Alignment(horizontal='left', vertical='top')

                    # Inhoud B20
                    c = 'B20'
                    ws[c].value = '0'
                    ws[c].alignment = Alignment(horizontal='center', vertical='top')

                    # Inhoud B21
                    c = 'B21'
                    ws[c].value = '1'
                    ws[c].alignment = Alignment(horizontal='center', vertical='top')

                    # Inhoud B22
                    c = 'B22'
                    ws[c].value = '2'
                    ws[c].alignment = Alignment(horizontal='center', vertical='top')

                    # Inhoud C6
                    c = 'C6'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C7
                    c = 'C7'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C9
                    c = 'C9'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C11
                    c = 'C11'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C13
                    c = 'C13'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C16
                    c = 'C16'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C20
                    c = 'C20'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C21
                    c = 'C21'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud C22
                    c = 'C22'
                    ws[c].value = '='
                    ws[c].alignment = Alignment(horizontal='center', vertical='center')

                    # Inhoud D6
                    c = 'D6'
                    ws[c].value = toelichting_rood
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D7
                    c = 'D7'
                    ws[c].value = toelichting_zwart
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D9
                    c = 'D9'
                    ws[c].value = toelichting_oranje
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D11
                    c = 'D11'
                    ws[c].value = toelichting_geel
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D13
                    c = 'D13'
                    ws[c].value = toelichting_groen
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D16
                    c = 'D16'
                    ws[c].value = toelichting_012
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D20
                    c = 'D20'
                    ws[c].value = toelichting_0
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D21
                    c = 'D21'
                    ws[c].value = toelichting_1
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Inhoud D22
                    c = 'D22'
                    ws[c].value = toelichting_2
                    ws[c].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                else:
                    # todo: (optioneel) opmaak toevoegen voor de kolomnamen
                    ws[f'A1'] = 'Job nummer'
                    ws[f'B1'] = 'Job omschrijving'
                    ws[f'C1'] = 'Uitvoerende'
                    ws[f'D1'] = 'PO'
                    ws[f'E1'] = 'Frequentie'
                    ws[f'F1'] = "Begin datum"
                    ws[f'G1'] = 'Eind datum'
                    ws[f'H1'] = 'SB'
                    ws[f'I1'] = 'HD'
                    ws[f'J1'] = 'TRA'
                    ws[f'K1'] = 'SSP'
                    ws[f'L1'] = 'Object'
                    ws[f'M1'] = 'Aannemer'
                    ws[f'N1'] = 'Wie'
                    ws[f'O1'] = 'Telefoonnummer'

            """
            Exporteren van data naar een excel bestand met de fuctie in/van pandas verwijderd de oorspronkelijke inhoud
            het bestand. 
            Openpyxl wordt nu gebruikt voor het schrijven van de data naar excelbestand
            """
            for i in range(len(te_lezen_weken)):
                # Bepalen van een active worksheet en het dataframe voor de juiste week
                if i == 0:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_1
                elif i == 1:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_2
                else:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_3

                # Schrijven van de data naar de sheets
                ws = export_wb.active
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)

            # Kleuren aan tabbladen toewijzen
            set_tab_color(export_wb)

            # Bericht voor de gebruiker
            print(f'Het bestand wordt opgeslagen')

            # Opslaan van het document
            export_wb.save(export_file_path)

        else:
            # Ophalen van het bestand
            export_wb = openpyxl.load_workbook(export_file_path)

            for i in range(len(te_lezen_weken)):
                # Bepalen van een active worksheet en het dataframe voor de juiste week
                if i == 0:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_1
                    kleuren = kleuren_week_1
                elif i == 1:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_2
                    kleuren = kleuren_week_2
                else:
                    export_wb.active = export_wb[f'Week {te_lezen_weken[i]}']
                    df = df_week_3
                    kleuren = kleuren_week_3

                ws = export_wb.active

                # Verwijderen van alle kleuren in het document
                for column in ws.columns:
                    for cell in column:
                        cell.fill = no_fill

                dubble_rows = []

                # Controleren of er gegevens aanwezig zijn in het tabblad
                if ws.max_row > 1:
                    # Controleren of job omschrijving in tabblad ook in nieuwe df staat
                    for ii in range(2, ws.max_row + 1):

                        job_omschrijving = ws[f'B{ii}']
                        uitvoerende = ws[f'C{ii}']
                        po = ws[f'D{ii}']
                        obj = ws[f'L{ii}']

                        for index, row in df.iterrows():

                            if df['Job omschrijving'][index] == job_omschrijving.value \
                                    and df['Uitvoerende'][index] == uitvoerende.value \
                                    and df['PO'][index] == po.value \
                                    and df['Object'][index] == obj.value:
                                dubble_rows.append(ii)
                            else:
                                pass

                # Verwijderen van de dubbele rijen
                for row_index in reversed(dubble_rows):
                    ws.delete_rows(row_index)

                # Oude unieke regels een kleur toewijzen
                for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = fill

                # De regels van het nieuwe df toevoegen aan het tabblad
                count = 0
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
                    if kleuren[count] != black:
                        ws[f'A{ws.max_row}'].fill = PatternFill(fill_type='solid',
                                                                start_color=kleuren[count],
                                                                end_color=kleuren[count])
                    else:
                        pass
                    ws.freeze_panes = 'A2'
                    count += 1

            # Kleuren aan tabbladen toewijzen
            set_tab_color(export_wb)

            # Bericht voor de gebruiker
            print(f'Het bestand wordt opgeslagen')

            # Opslaan van het document
            export_wb.save(export_file_path)

    print("Het programma is afgerond. U kunt de planning ophalen.")
